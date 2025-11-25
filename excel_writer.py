
from openpyxl import load_workbook
import sys

class ExcelWriter:
    def __init__(self, file_path: str, sheet_name: str):
        try:
            self.file_path = file_path
            self.wb = load_workbook(file_path)
            if sheet_name not in self.wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
            self.sheet = self.wb[sheet_name]
        except Exception as e:
            print(f"Error opening workbook: {e}")
            sys.exit(1)

    def write_row(self, df, row: int, data: dict, columns: list):
        """
        Write data to specified columns in the given row.
        :param df: pandas DataFrame (for column index mapping)
        :param row: Row number (1-based)
        :param data: Dictionary of column -> value
        :param columns: List of columns to write
        """
        col_indexes = df.columns
        for col in columns:
            if col not in col_indexes:
                raise KeyError(f"Column '{col}' not found in DataFrame.")
            col_index = col_indexes.get_loc(col) + 1  # openpyxl is 1-based
            self.sheet.cell(row=row, column=col_index).value = data.get(col)
            self.wb.save(self.file_path)

    def save_and_close(self):
        try:
            self.wb.save(self.file_path)
            self.wb.close()
        except PermissionError:
            print(f"Cannot write to {self.file_path} — file is open in Excel or in use.")
            sys.exit(1)
        except Exception as e:
            print(f"Unexpected error while saving: {e}")
            sys.exit(1)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.save_and_close()
